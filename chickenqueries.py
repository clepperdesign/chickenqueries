from openpyxl import load_workbook
wb = load_workbook('ch-database.xlsx')
ws = wb.active
def sortbymultiple():
    mult = str.lower(input("Enter filter criteria: Type, Free Range, Cold Hardy, and/or Heat Hardy \n"))
    endlist = []
    frlist = []
    heatlist = []
    coldlist = []
    def chxtype():
        typesort = str.lower(input("Enter desired type: Eggs, Meat, or Dual \n"))
        for row in ws.iter_rows():
            for cell in row:
                if 'egg' in typesort:
                    if (cell.value == 'eggs' or cell.value == 'dual') and len(endlist)==0:
                        print(row[0].value + ': ' + str(row[2].value) + ' eggs per year')
                    elif cell.value == 'eggs' or cell.value == 'dual':
                        if row[0].value in endlist:
                            print(row[0].value + ': ' + str(row[2].value) + ' eggs per year')
                elif 'meat' in typesort:
                    if (cell.value == 'meat' or cell.value == 'dual') and len(endlist)==0:
                        print(row[0].value + ': weight from ' + str(row[3].value) + ' to ' + str(row[4].value) + ' pounds.')
                    elif cell.value == 'meat' or cell.value == 'dual':
                        if row[0].value in endlist:
                            print(row[0].value + ': weight from ' + str(row[3].value) + ' to ' + str(row[4].value) + ' pounds.')
                elif 'dual' in typesort:
                    if cell.value == 'dual' and len(endlist)==0:
                        print(row[0].value + ': ' + str(row[2].value) + ' eggs per year and weight from ' + str(row[3].value) + ' to ' + str(row[4].value) + ' pounds.')
                    elif cell.value == 'dual':
                        print(row[0].value + ': ' + str(row[2].value) + ' eggs per year and weight from ' + str(row[3].value) + ' to ' + str(row[4].value) + ' pounds.')
                else:
                    print('cluck')
    def freerange():
        for row in ws.iter_rows():
            if row[5].value == True:
                    frlist.append(row[0].value)
    def cold():
        for row in ws.iter_rows():
            if row[6].value == True:
                coldlist.append(row[0].value)
    def hot():
        for row in ws.iter_rows():
            if row[7].value == True:
                heatlist.append(row[0].value)
    def endlist_compile():
        for item in frlist:
            if len(coldlist)==0 and len(heatlist)==0:
                endlist.append(item)
            elif len(coldlist)==0:
                if item in heatlist:
                    endlist.append(item)
            elif len(heatlist)==0:
                if item in coldlist:
                    endlist.append(item)
            elif item in coldlist and item in heatlist:
                endlist.append(item)
        for item in heatlist:
            if len(coldlist)==0 and len(frlist)==0:
                endlist.append(item)
            elif len(coldlist)==0:
                if item in frlist:
                    endlist.append(item)
            elif len(frlist)==0:
                if item in coldlist:
                    endlist.append(item)
        for item in coldlist:
            if len(heatlist)==0 and len(frlist)==0:
                endlist.append(item)
            elif len(heatlist)==0:
                if item in frlist:
                    endlist.append(item)
            elif len(frlist)==0:
                if item in heatlist and item not in endlist:
                    endlist.append(item)
    def endlist_print():
        for item in endlist:
            print(item)
    if 'free' in mult:
        freerange()
    if 'cold' in mult:
        cold()
    if 'heat' in mult:
        hot()
    endlist_compile()
    if 'type' in mult:
        chxtype()
    elif 'type' not in mult:
      endlist_print()  
sortbymultiple()
